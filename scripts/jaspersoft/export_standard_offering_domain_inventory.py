#!/usr/bin/env python3
"""
Export Standard Offering Jaspersoft domain inventory from a tenant-root import ZIP.

Produces:
- Excel workbook with one sheet per domain (+ Index)
- Master CSV files for domains, fields, joins, and tables
- Per-domain CSV folder
"""

from __future__ import annotations

import argparse
import csv
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

NS = {"sl": "http://www.jaspersoft.com/2007/SL/XMLSchema"}

WORKSTREAM_LABELS = {
    "Billing_and_Rates": "Billing and Rates",
    "Cashiering": "Cashiering",
    "Common": "Common",
    "Customer_Operations": "Customer Operations",
    "Debt_Management": "Debt Management",
    "Development": "Development",
    "Field_Operations": "Field Operations",
    "Finance": "Finance",
    "Meter_Operations": "Meter Operations",
    "New_Services___Planning": "New Services and Planning",
}


@dataclass
class DomainRecord:
    domain_name: str
    domain_label: str
    description: str
    repository_path: str
    workstream_folder: str
    workstream_label: str
    subfolder: str
    datasource_alias: str
    tables: list[dict[str, str]] = field(default_factory=list)
    joins: list[dict[str, str]] = field(default_factory=list)
    fields: list[dict[str, str]] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export domain inventory from import ZIP.")
    parser.add_argument(
        "--source-zip",
        default="deploy/jaspersoft_environment_promotion/prepared_imports/standard_offering_Origin_DEMO_import.zip",
        help="Tenant-root Standard Offering import ZIP.",
    )
    parser.add_argument(
        "--output-dir",
        default="output/standard_offering_domain_inventory",
        help="Directory for workbook and CSV outputs.",
    )
    return parser.parse_args()


def _text(element: ET.Element | None, tag: str) -> str:
    if element is None:
        return ""
    child = element.find(tag)
    return (child.text or "").strip() if child is not None else ""


def parse_domain_xml(xml_text: str, repo_rel: str) -> dict[str, str]:
    root = ET.fromstring(xml_text)
    folder = _text(root, "folder")
    parts = [p for p in folder.split("/") if p]
    workstream_folder = ""
    subfolder = ""
    if "Standard_Offering" in parts:
        idx = parts.index("Standard_Offering")
        if idx + 1 < len(parts):
            workstream_folder = parts[idx + 1]
        if idx + 2 < len(parts):
            subfolder = "/".join(parts[idx + 2:])
    elif repo_rel.startswith("resources/SmartCity/Report/Standard_Offering/"):
        rel = repo_rel.split("Standard_Offering/")[-1]
        rel_parts = rel.split("/")
        if len(rel_parts) >= 2:
            workstream_folder = rel_parts[0]
            subfolder = "/".join(rel_parts[1:-1]) if len(rel_parts) > 2 else rel_parts[0]

    if not workstream_folder and repo_rel:
        m = re.search(r"Standard_Offering/([^/]+)/", repo_rel)
        if m:
            workstream_folder = m.group(1)
        m2 = re.search(r"Standard_Offering/([^/]+/.*?)/[^/]+___Domain", repo_rel)
        if m2:
            subfolder = m2.group(1).split("/", 1)[-1] if "/" in m2.group(1) else ""

    return {
        "domain_name": _text(root, "name"),
        "domain_label": _text(root, "label"),
        "description": _text(root, "description"),
        "folder": folder,
        "repository_path": folder or repo_rel,
        "workstream_folder": workstream_folder,
        "workstream_label": WORKSTREAM_LABELS.get(workstream_folder, workstream_folder.replace("_", " ")),
        "subfolder": subfolder.replace("___Domain.xml", "").replace(".xml", ""),
        "datasource_alias": _text(root.find("dataSource"), "alias") if root.find("dataSource") is not None else "",
    }


def parse_schema(schema_text: str) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    root = ET.fromstring(schema_text)
    parent_map = {child: parent for parent in root.iter() for child in parent}

    tables: list[dict[str, str]] = []
    seen_table_ids: set[str] = set()
    for table_el in root.findall(".//sl:jdbcTable", NS):
        table_id = table_el.get("id") or ""
        if table_id in seen_table_ids:
            continue
        seen_table_ids.add(table_id)
        join_list = table_el.find("sl:joinList", NS)
        is_join_tree = join_list is not None
        tables.append(
            {
                "domain_table_id": table_id,
                "physical_table": table_el.get("datasourceTableName") or "",
                "schema": table_el.get("schemaAlias") or "CISADM",
                "role": "join_tree_root" if is_join_tree else "physical_table",
            }
        )

    joins: list[dict[str, str]] = []
    join_seq = 0
    for join_list in root.findall(".//sl:joinList", NS):
        parent = parent_map.get(join_list)
        join_tree_id = parent.get("id") if parent is not None else ""
        for join_el in join_list.findall("sl:join", NS):
            join_seq += 1
            joins.append(
                {
                    "join_tree_id": join_tree_id,
                    "join_sequence": str(join_seq),
                    "join_type": join_el.get("type") or "",
                    "left_table": join_el.get("left") or "",
                    "right_table": join_el.get("right") or "",
                    "join_expression": join_el.get("expr") or "",
                }
            )

    if not joins and tables:
        primary = next((t for t in tables if t["role"] == "join_tree_root"), tables[0])
        joins.append(
            {
                "join_tree_id": primary["domain_table_id"],
                "join_sequence": "1",
                "join_type": "single_table",
                "left_table": primary["domain_table_id"],
                "right_table": "",
                "join_expression": f"Single-table domain on {primary['physical_table']}",
            }
        )

    fields: list[dict[str, str]] = []
    for item_group in root.findall(".//sl:itemGroup", NS):
        ig_id = item_group.get("id") or ""
        ig_label = item_group.get("label") or ""
        ig_resource = item_group.get("resourceId") or ""
        items_parent = item_group.find("sl:items", NS)
        if items_parent is None:
            continue
        for item in items_parent.findall("sl:item", NS):
            resource_id = item.get("resourceId") or ""
            source_table = ""
            source_column = ""
            if "." in resource_id:
                source_table, source_column = resource_id.rsplit(".", 1)
            fields.append(
                {
                    "item_group_id": ig_id,
                    "item_group_label": ig_label,
                    "item_group_resource": ig_resource,
                    "field_id": item.get("id") or "",
                    "field_label": item.get("label") or "",
                    "resource_id": resource_id,
                    "source_table": source_table,
                    "source_column": source_column,
                    "dimension_or_measure": item.get("dimensionOrMeasure") or "",
                    "default_aggregation": item.get("defaultAgg") or "",
                }
            )

    return tables, joins, fields


def load_domains_from_zip(zip_path: Path) -> list[DomainRecord]:
    domains: list[DomainRecord] = []
    with zipfile.ZipFile(zip_path) as archive:
        domain_xml_paths = sorted(
            name
            for name in archive.namelist()
            if "Standard_Offering/" in name
            and name.endswith("___Domain.xml")
            and "/_files/" not in name
        )
        for domain_xml_path in domain_xml_paths:
            schema_path = domain_xml_path.replace(".xml", "_files/schema.data")
            if schema_path not in archive.namelist():
                continue
            meta = parse_domain_xml(
                archive.read(domain_xml_path).decode("utf-8", errors="replace"),
                domain_xml_path,
            )
            rel_path = domain_xml_path.split("Standard_Offering/")[-1]
            if not meta["subfolder"]:
                parts = rel_path.split("/")
                meta["subfolder"] = "/".join(parts[:-1]) if len(parts) > 1 else ""

            tables, joins, fields = parse_schema(
                archive.read(schema_path).decode("utf-8", errors="replace")
            )
            domains.append(
                DomainRecord(
                    domain_name=meta["domain_name"],
                    domain_label=meta["domain_label"],
                    description=meta["description"],
                    repository_path=f"/SmartCity/Report/Standard_Offering/{rel_path.replace('.xml', '')}",
                    workstream_folder=meta["workstream_folder"],
                    workstream_label=meta["workstream_label"],
                    subfolder=meta["subfolder"],
                    datasource_alias=meta["datasource_alias"],
                    tables=tables,
                    joins=joins,
                    fields=fields,
                )
            )
    domains.sort(key=lambda d: (d.workstream_folder, d.domain_name))
    return domains


def safe_sheet_name(domain_name: str, used: set[str]) -> str:
    base = re.sub(r"[^\w\- ]", "_", domain_name.replace("___Domain", ""))[:28]
    name = base or "Domain"
    counter = 1
    candidate = name
    while candidate in used:
        suffix = f"_{counter}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        counter += 1
    used.add(candidate)
    return candidate


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_domain_sheet(ws, domain: DomainRecord) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    row = 1
    ws.cell(row, 1, "Domain Name").font = bold
    ws.cell(row, 2, domain.domain_name)
    row += 1
    ws.cell(row, 1, "Domain Label").font = bold
    ws.cell(row, 2, domain.domain_label)
    row += 1
    ws.cell(row, 1, "Workstream").font = bold
    ws.cell(row, 2, domain.workstream_label)
    row += 1
    ws.cell(row, 1, "Workstream Folder").font = bold
    ws.cell(row, 2, domain.workstream_folder)
    row += 1
    ws.cell(row, 1, "Subfolder").font = bold
    ws.cell(row, 2, domain.subfolder)
    row += 2

    def write_section(title: str, headers: list[str], rows: list[dict[str, str]], keys: list[str]) -> int:
        nonlocal row
        ws.cell(row, 1, title).font = bold
        row += 1
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = bold
        row += 1
        for data in rows:
            for col, key in enumerate(keys, start=1):
                ws.cell(row, col, data.get(key, ""))
            row += 1
        row += 1
        return row

    write_section(
        "Physical Tables",
        ["Domain Table Alias", "Physical Table"],
        domain.tables,
        ["domain_table_id", "physical_table"],
    )
    write_section(
        "Join Tree",
        ["Join Tree", "Sequence", "Join Type", "Left Table", "Right Table", "Join Expression"],
        domain.joins,
        ["join_tree_id", "join_sequence", "join_type", "left_table", "right_table", "join_expression"],
    )
    write_section(
        "Exposed Fields",
        [
            "Item Group",
            "Field Label",
            "Source Table",
            "Source Column",
            "Dimension/Measure",
            "Default Aggregation",
        ],
        domain.fields,
        [
            "item_group_label",
            "field_label",
            "source_table",
            "source_column",
            "dimension_or_measure",
            "default_aggregation",
        ],
    )

    for col in range(1, 10):
        letter = get_column_letter(col)
        max_len = 12
        for cell in ws[letter]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[letter].width = max_len + 2


def export_workbook(domains: list[DomainRecord], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    index = wb.active
    index.title = "Index"
    headers = [
        "Domain Name",
        "Domain Label",
        "Workstream",
        "Subfolder",
        "Table Count",
        "Join Count",
        "Field Count",
        "Sheet Name",
    ]
    bold = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        index.cell(1, col, header).font = bold

    used_sheet_names: set[str] = {"Index"}
    for row_idx, domain in enumerate(domains, start=2):
        sheet_name = safe_sheet_name(domain.domain_name, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        write_domain_sheet(ws, domain)
        values = [
            domain.domain_name,
            domain.domain_label,
            domain.workstream_label,
            domain.subfolder,
            len(domain.tables),
            len(domain.joins),
            len(domain.fields),
            sheet_name,
        ]
        for col, value in enumerate(values, start=1):
            index.cell(row_idx, col, value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[2]
    zip_path = (repo_root / args.source_zip).resolve()
    output_dir = (repo_root / args.output_dir).resolve()

    if not zip_path.exists():
        raise SystemExit(f"Source ZIP not found: {zip_path}")

    domains = load_domains_from_zip(zip_path)
    if not domains:
        raise SystemExit("No domains found in source ZIP.")

    per_domain_dir = output_dir / "by_domain"
    per_domain_dir.mkdir(parents=True, exist_ok=True)

    index_rows: list[dict[str, Any]] = []
    table_rows: list[dict[str, Any]] = []
    join_rows: list[dict[str, Any]] = []
    field_rows: list[dict[str, Any]] = []

    for domain in domains:
        index_rows.append(
            {
                "domain_name": domain.domain_name,
                "domain_label": domain.domain_label,
                "workstream": domain.workstream_label,
                "workstream_folder": domain.workstream_folder,
                "subfolder": domain.subfolder,
                "table_count": len(domain.tables),
                "join_count": len(domain.joins),
                "field_count": len(domain.fields),
            }
        )
        slug = re.sub(
            r"[^\w\-]+",
            "_",
            f"{domain.workstream_folder}_{domain.subfolder}_{domain.domain_name}".replace("___Domain", ""),
        ).strip("_")
        slug = re.sub(r"_+", "_", slug)[:120]
        domain_csv = per_domain_dir / f"{slug}.csv"

        with domain_csv.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["section", "domain_name", domain.domain_name])
            writer.writerow(["section", "workstream", domain.workstream_label])
            writer.writerow([])
            writer.writerow(["PHYSICAL_TABLES"])
            writer.writerow(["domain_table_alias", "physical_table"])
            for t in domain.tables:
                writer.writerow([t["domain_table_id"], t["physical_table"]])
            writer.writerow([])
            writer.writerow(["JOIN_TREE"])
            writer.writerow(
                ["join_tree", "join_sequence", "join_type", "left_table", "right_table", "join_expression"]
            )
            for j in domain.joins:
                writer.writerow(
                    [
                        j["join_tree_id"],
                        j["join_sequence"],
                        j["join_type"],
                        j["left_table"],
                        j["right_table"],
                        j["join_expression"],
                    ]
                )
            writer.writerow([])
            writer.writerow(["EXPOSED_FIELDS"])
            writer.writerow(
                [
                    "item_group_label",
                    "field_label",
                    "source_table",
                    "source_column",
                    "dimension_or_measure",
                    "default_aggregation",
                ]
            )
            for f in domain.fields:
                writer.writerow(
                    [
                        f["item_group_label"],
                        f["field_label"],
                        f["source_table"],
                        f["source_column"],
                        f["dimension_or_measure"],
                        f["default_aggregation"],
                    ]
                )

        for t in domain.tables:
            table_rows.append(
                {
                    "domain_name": domain.domain_name,
                    "workstream": domain.workstream_label,
                    "domain_table_alias": t["domain_table_id"],
                    "physical_table": t["physical_table"],
                }
            )
        for j in domain.joins:
            join_rows.append({"domain_name": domain.domain_name, "workstream": domain.workstream_label, **j})
        for f in domain.fields:
            field_rows.append(
                {
                    "domain_name": domain.domain_name,
                    "workstream": domain.workstream_label,
                    "item_group_label": f["item_group_label"],
                    "field_label": f["field_label"],
                    "source_table": f["source_table"],
                    "source_column": f["source_column"],
                    "dimension_or_measure": f["dimension_or_measure"],
                    "default_aggregation": f["default_aggregation"],
                }
            )

    write_csv(
        output_dir / "domain_index.csv",
        index_rows,
        list(index_rows[0].keys()),
    )
    write_csv(
        output_dir / "domain_tables_master.csv",
        table_rows,
        ["domain_name", "workstream", "domain_table_alias", "physical_table"],
    )
    write_csv(
        output_dir / "domain_joins_master.csv",
        join_rows,
        [
            "domain_name",
            "workstream",
            "join_tree_id",
            "join_sequence",
            "join_type",
            "left_table",
            "right_table",
            "join_expression",
        ],
    )
    write_csv(
        output_dir / "domain_fields_master.csv",
        field_rows,
        [
            "domain_name",
            "workstream",
            "item_group_label",
            "field_label",
            "source_table",
            "source_column",
            "dimension_or_measure",
            "default_aggregation",
        ],
    )

    workbook_path = output_dir / "standard_offering_domain_inventory.xlsx"
    export_workbook(domains, workbook_path)

    print(f"Domains exported: {len(domains)}")
    print(f"Workbook: {workbook_path}")
    print(f"Index CSV: {output_dir / 'domain_index.csv'}")
    print(f"Fields master CSV: {output_dir / 'domain_fields_master.csv'}")
    print(f"Per-domain CSV folder: {per_domain_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
