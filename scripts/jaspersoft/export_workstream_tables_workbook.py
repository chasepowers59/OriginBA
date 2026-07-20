#!/usr/bin/env python3
"""
Export workstream-level Excel workbook: tables, join keys, and table content summaries.

Built from Standard Offering domain inventory (import ZIP) plus the workstream
reporting dictionary for CISADM table descriptions.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import sys

_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from export_standard_offering_domain_inventory import (
    WORKSTREAM_LABELS,
    DomainRecord,
    load_domains_from_zip,
)


SNAPSHOT_TABLES: dict[str, dict[str, str]] = {
    "BSEG_BILLED_USAGE_RPT_CURR": {
        "grain": "One row per BSEG_ID",
        "summary": "Billed usage and billed amount at completed bill-segment grain.",
    },
    "BSEG_SQ_USAGE_RPT_CURR": {
        "grain": "One row per bill-segment determinant key",
        "summary": "Billed quantity by UOM / TOU / SQI at determinant grain.",
    },
    "FT_RPT_CURR": {
        "grain": "One row per FT_ID",
        "summary": "Financial transaction header amounts, status, and trace context.",
    },
    "FT_GL_DISTRIBUTION_RPT_CURR": {
        "grain": "One row per FT_ID and GL_SEQ_NBR",
        "summary": "GL account and distribution-code detail for FT lines.",
    },
    "ACCT_DEBT_RPT_CURR": {
        "grain": "One row per ACCT_ID",
        "summary": "Account-level debt exposure and aging buckets.",
    },
    "COLL_PROC_RPT_CURR": {
        "grain": "One row per COLL_PROC_ID",
        "summary": "Collection process workflow and status monitoring.",
    },
    "D1_USAGE_RPT_CURR": {
        "grain": "One row per D1_USAGE_ID",
        "summary": "Usage process header monitoring.",
    },
    "D1_USAGE_SCALAR_DTL_RPT_CURR": {
        "grain": "One row per D1_USAGE_ID and SEQ_NUM",
        "summary": "Scalar quantity detail for usage transactions.",
    },
    "D1_MSRMT_RPT_CURR": {
        "grain": "One row per final processed measurement",
        "summary": "Final measurement values and VEE context.",
    },
    "PAY_TNDR_CASH_RPT_CURR": {
        "grain": "One row per PAY_TENDER_ID",
        "summary": "Payment intake, tender type, and cashiering totals.",
    },
}

WORKSTREAM_ORDER = [
    "Billing_and_Rates",
    "Cashiering",
    "Common",
    "Customer_Operations",
    "Debt_Management",
    "Field_Operations",
    "Finance",
    "Meter_Operations",
    "New_Services___Planning",
]


@dataclass
class TableSummary:
    physical_table: str
    table_type: str
    description: str
    domains: set[str] = field(default_factory=set)
    item_groups: set[str] = field(default_factory=set)
    field_labels: list[str] = field(default_factory=list)
    join_keys: set[str] = field(default_factory=set)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export workstream tables and join keys workbook."
    )
    parser.add_argument(
        "--source-zip",
        default="deploy/jaspersoft_environment_promotion/prepared_imports/standard_offering_Origin_DEMO_import.zip",
        help="Tenant-root Standard Offering import ZIP.",
    )
    parser.add_argument(
        "--dictionary-json",
        default="output/workstream_reporting_dictionary.json",
        help="Workstream reporting dictionary for table descriptions.",
    )
    parser.add_argument(
        "--output",
        default="output/standard_offering_domain_inventory/workstream_tables_and_joins.xlsx",
        help="Output Excel workbook path.",
    )
    return parser.parse_args()


def load_table_descriptions(dictionary_path: Path) -> dict[str, str]:
    if not dictionary_path.exists():
        return {}
    data = json.loads(dictionary_path.read_text(encoding="utf-8"))
    descriptions: dict[str, str] = {}
    for workstream in data.values():
        for table_name, meta in workstream.get("tables", {}).items():
            desc = (meta.get("description") or "").strip()
            if table_name not in descriptions and desc:
                descriptions[table_name] = desc
    return descriptions


def alias_maps(domain: DomainRecord) -> dict[str, str]:
    mapping = {t["domain_table_id"]: t["physical_table"] for t in domain.tables}
    for table in domain.tables:
        mapping.setdefault(table["physical_table"], table["physical_table"])
    return mapping


def resolve_physical_table(source_table: str, mapping: dict[str, str]) -> str:
    if not source_table:
        return ""
    if "." in source_table:
        alias, table_name = source_table.rsplit(".", 1)
        if alias in mapping:
            return mapping[alias]
        return table_name
    return mapping.get(source_table, source_table)


def table_type_for(physical_table: str) -> str:
    if physical_table in SNAPSHOT_TABLES or physical_table.endswith("_RPT_CURR"):
        return "Governed Snapshot"
    if physical_table.startswith("CI_"):
        return "CISADM Base"
    if physical_table.startswith("D1_"):
        return "C2M Base"
    if physical_table.startswith("F1_"):
        return "Framework Base"
    if physical_table.startswith("W1_"):
        return "Work Asset Base"
    if physical_table.startswith("SC_"):
        return "Security / Reference"
    return "Reference / Other"


def table_description(physical_table: str, descriptions: dict[str, str]) -> str:
    if physical_table in SNAPSHOT_TABLES:
        snap = SNAPSHOT_TABLES[physical_table]
        return f"{snap['summary']} Grain: {snap['grain']}."
    return descriptions.get(physical_table, "")


def extract_join_key_columns(physical_table: str, join_expression: str) -> list[str]:
    if not join_expression or join_expression.startswith("Single-table"):
        return []
    pattern = re.compile(
        rf"{re.escape(physical_table)}\.([A-Z0-9_]+)",
        flags=re.IGNORECASE,
    )
    return sorted(set(pattern.findall(join_expression)))


def summarize_workstream(
    domains: list[DomainRecord],
    descriptions: dict[str, str],
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    domain_rows: list[dict[str, str]] = []
    tables: dict[str, TableSummary] = {}
    join_rows: list[dict[str, str]] = []

    for domain in domains:
        mapping = alias_maps(domain)
        domain_rows.append(
            {
                "domain_label": domain.domain_label,
                "subfolder": domain.subfolder,
                "table_count": str(len(domain.tables)),
                "join_count": str(len(domain.joins)),
                "field_count": str(len(domain.fields)),
            }
        )

        for join in domain.joins:
            if join["join_type"] == "single_table":
                continue
            join_rows.append(
                {
                    "domain_label": domain.domain_label,
                    "join_sequence": join["join_sequence"],
                    "join_type": join["join_type"],
                    "left_table": join["left_table"],
                    "right_table": join["right_table"],
                    "join_expression": join["join_expression"],
                }
            )

        physical_tables_in_domain = {
            t["physical_table"] for t in domain.tables if t["physical_table"]
        }
        for physical_table in physical_tables_in_domain:
            summary = tables.setdefault(
                physical_table,
                TableSummary(
                    physical_table=physical_table,
                    table_type=table_type_for(physical_table),
                    description=table_description(physical_table, descriptions),
                ),
            )
            summary.domains.add(domain.domain_label)

        for join in domain.joins:
            expr = join["join_expression"]
            for side in (join["left_table"], join["right_table"]):
                physical = mapping.get(side, side)
                if physical not in tables:
                    continue
                for column in extract_join_key_columns(physical, expr):
                    tables[physical].join_keys.add(column)

        seen_labels: dict[str, set[str]] = defaultdict(set)
        for field_row in domain.fields:
            physical = resolve_physical_table(field_row["source_table"], mapping)
            if not physical:
                continue
            summary = tables.setdefault(
                physical,
                TableSummary(
                    physical_table=physical,
                    table_type=table_type_for(physical),
                    description=table_description(physical, descriptions),
                ),
            )
            summary.domains.add(domain.domain_label)
            group = field_row["item_group_label"]
            label = field_row["field_label"]
            if group:
                summary.item_groups.add(group)
            if label and label not in seen_labels[physical]:
                seen_labels[physical].add(label)
                if len(summary.field_labels) < 12:
                    summary.field_labels.append(label)

    table_rows: list[dict[str, str]] = []
    for physical_table in sorted(tables):
        summary = tables[physical_table]
        table_rows.append(
            {
                "physical_table": physical_table,
                "table_type": summary.table_type,
                "description": summary.description,
                "used_in_domains": "; ".join(sorted(summary.domains)),
                "typical_content": "; ".join(sorted(summary.item_groups)),
                "sample_fields": "; ".join(summary.field_labels),
                "common_join_keys": "; ".join(sorted(summary.join_keys)),
            }
        )

    return domain_rows, table_rows, join_rows


def write_section(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[dict[str, str]],
    keys: list[str],
    bold,
) -> int:
    row = start_row
    ws.cell(row, 1, title).font = bold
    row += 1
    for col, header in enumerate(headers, start=1):
        ws.cell(row, col, header).font = bold
    row += 1
    for data in rows:
        for col, key in enumerate(keys, start=1):
            ws.cell(row, col, data.get(key, ""))
        row += 1
    return row + 1


def autosize_columns(ws, max_col: int) -> None:
    from openpyxl.utils import get_column_letter

    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 12
        for cell in ws[letter]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 90))
        ws.column_dimensions[letter].width = max_len + 2


def export_workbook(
    grouped_domains: dict[str, list[DomainRecord]],
    descriptions: dict[str, str],
    output_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    index = wb.active
    index.title = "Index"
    bold = Font(bold=True)
    index_headers = [
        "Workstream",
        "Domain Count",
        "Distinct Tables",
        "Join Rows",
        "Sheet Name",
    ]
    for col, header in enumerate(index_headers, start=1):
        index.cell(1, col, header).font = bold

    used_sheet_names: set[str] = {"Index"}
    index_row = 2

    for workstream_folder in WORKSTREAM_ORDER:
        domains = grouped_domains.get(workstream_folder, [])
        if not domains:
            continue

        label = WORKSTREAM_LABELS.get(
            workstream_folder, workstream_folder.replace("_", " ")
        )
        sheet_name = label[:31]
        suffix = 1
        while sheet_name in used_sheet_names:
            sheet_name = f"{label[:28]}_{suffix}"
            suffix += 1
        used_sheet_names.add(sheet_name)

        domain_rows, table_rows, join_rows = summarize_workstream(domains, descriptions)
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(1, 1, "Workstream").font = bold
        ws.cell(1, 2, label)
        ws.cell(2, 1, "Standard Offering Domains").font = bold
        ws.cell(2, 2, str(len(domains)))
        row = 4

        row = write_section(
            ws,
            row,
            "Domains In This Workstream",
            ["Domain Label", "Subfolder", "Tables", "Joins", "Fields"],
            domain_rows,
            ["domain_label", "subfolder", "table_count", "join_count", "field_count"],
            bold,
        )
        row = write_section(
            ws,
            row,
            "Physical Tables",
            [
                "Physical Table",
                "Table Type",
                "Description / Grain",
                "Used In Domains",
                "Typical Content (Item Groups)",
                "Sample Fields",
                "Common Join Keys",
            ],
            table_rows,
            [
                "physical_table",
                "table_type",
                "description",
                "used_in_domains",
                "typical_content",
                "sample_fields",
                "common_join_keys",
            ],
            bold,
        )
        write_section(
            ws,
            row,
            "Join Keys By Domain",
            [
                "Domain",
                "Sequence",
                "Join Type",
                "Left Table",
                "Right Table",
                "Join Expression",
            ],
            join_rows,
            [
                "domain_label",
                "join_sequence",
                "join_type",
                "left_table",
                "right_table",
                "join_expression",
            ],
            bold,
        )
        autosize_columns(ws, 7)

        index.cell(index_row, 1, label)
        index.cell(index_row, 2, len(domains))
        index.cell(index_row, 3, len(table_rows))
        index.cell(index_row, 4, len(join_rows))
        index.cell(index_row, 5, sheet_name)
        index_row += 1

    autosize_columns(index, 5)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[2]
    zip_path = (repo_root / args.source_zip).resolve()
    dictionary_path = (repo_root / args.dictionary_json).resolve()
    output_path = (repo_root / args.output).resolve()

    if not zip_path.exists():
        raise SystemExit(f"Source ZIP not found: {zip_path}")

    domains = load_domains_from_zip(zip_path)
    if not domains:
        raise SystemExit("No domains found in source ZIP.")

    grouped: dict[str, list[DomainRecord]] = defaultdict(list)
    for domain in domains:
        grouped[domain.workstream_folder].append(domain)

    descriptions = load_table_descriptions(dictionary_path)
    export_workbook(grouped, descriptions, output_path)

    print(f"Workbook: {output_path}")
    print(f"Workstreams: {len([k for k in WORKSTREAM_ORDER if grouped.get(k)])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
